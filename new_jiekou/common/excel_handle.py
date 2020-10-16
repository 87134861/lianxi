"""
AUthor:regan
email:8713*****@qq.com
company:wuhan****
"""
import openpyxl

# # 读取excel文件，返回一个workbook对象
# wb=openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\ligen.xlsx")
# # 获取工作簿中的表单
# sh=wb["Sheet1"]
# # 获取第一行第一列的内容 row代表行，column代表列
# data=sh.cell(row=1,column=1).value
# print(data)
# 写入文件
# wb2=openpyxl.load_workbook("ligen.xlsx")
# wb2["Sheet1"].cell(row=1,column=1,value="hello")
# wb2.save("ligen.xlsx")

# # 读取文件中的行
# wb3=openpyxl.load_workbook("ligen.xlsx")
# kk=list(wb3["Sheet1"].rows)
# print(kk[1])
# # # 读取指定行的数据
# # li=[]
# # for i in kk[2]:
# #     li.append(i.value)
# # print(li)
# # 通过列表推导式获取
# li1=[i.value for i in kk[2]]
# print(li1)
# 获取指定列的数据
# wb4=openpyxl.load_workbook("ligen.xlsx")
# mm=list(wb4["Sheet1"].columns)
# li2=[]
# for i in mm[1]:
#     li2.append(i.value)
# print(li2)
# print(wb4["Sheet1"].max_row)
# print(wb4["Sheet1"].max_column)


class Excel_Read:
    def __init__(self,filename,Sheetname,n,m):
        self.filename=filename
        self.Sheetname=Sheetname
        self.n=n
        self.m=m

    def excel_read(self):
        # filename=r"C:\Users\Administrator\Desktop\ligen.xlsx"
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.Sheetname]
        data = sh.cell(row=self.n, column=self.m).value
        return data
#
# aa=excel_read(r"C:\Users\Administrator\Desktop\ligen.xlsx","Sheet1",1,1)
#
# bb=aa.excel_read()
# print(bb)
# print(eval(bb))
# aa="{'name':'ligen','age':'32'}"
# print(eval(aa))
# data=Excel_Read(r"D:\new_jiekou\data\api.xlsx","Sheet1",2,6)
# bb=data.excel_read()
# print(eval(bb))
